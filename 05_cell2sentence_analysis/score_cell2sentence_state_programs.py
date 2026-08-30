from __future__ import annotations

import argparse
import csv
import json
import os
import statistics
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


OUTDIR: Path
META: Path
PRED: Path
MPL_CACHE: Path


ASTRO_MARKERS = {
    "astro_homeostatic": [
        "ALDH1L1",
        "AQP4",
        "SLC1A2",
        "SLC1A3",
        "GLUL",
        "GJA1",
        "S100B",
        "SOX9",
        "GFAP",
        "APOE",
        "FGFR3",
        "CLU",
        "SPARCL1",
        "FABP7",
    ],
    "astro_reactive_active": [
        "GFAP",
        "VIM",
        "CD44",
        "C3",
        "SERPINA3",
        "LCN2",
        "CHI3L1",
        "OSMR",
        "STAT3",
        "SOCS3",
        "JUN",
        "FOS",
        "CEBPB",
        "NFKBIA",
        "ICAM1",
        "VCAM1",
        "PTX3",
        "CXCL10",
        "CCL2",
        "IL6",
        "IFITM3",
        "SERPING1",
        "HLA-DRA",
        "HLA-DRB1",
        "CD74",
        "TIMP1",
        "S100A10",
        "ANXA1",
    ],
    "astro_inflammatory_neurotoxic": [
        "C3",
        "SERPING1",
        "GBP2",
        "MX1",
        "IFITM3",
        "STAT1",
        "CXCL10",
        "CCL2",
        "IL1B",
        "TNF",
        "NFKBIA",
        "IRF1",
        "HLA-DRA",
        "CD74",
    ],
    "astro_reparative_trophic": [
        "S100A10",
        "PTX3",
        "EMP1",
        "SPHK1",
        "CD109",
        "CLCF1",
        "TGM1",
        "TM4SF1",
        "SPARCL1",
        "CLU",
        "APOE",
        "THBS1",
        "VEGFA",
    ],
}

MICROGLIA_MARKERS = {
    "microglia_homeostatic": [
        "P2RY12",
        "TMEM119",
        "SALL1",
        "HEXB",
        "CX3CR1",
        "GPR34",
        "SLC2A5",
        "BIN1",
        "CSF1R",
        "TREM2",
    ],
    "microglia_m1_inflammatory": [
        "IL1B",
        "TNF",
        "IL6",
        "CXCL9",
        "CXCL10",
        "CXCL11",
        "CCL2",
        "CCL3",
        "CCL4",
        "NFKBIA",
        "STAT1",
        "IRF1",
        "NOS2",
        "CD86",
    ],
    "microglia_m2_immunoregulatory": [
        "MRC1",
        "CD163",
        "MSR1",
        "IL10",
        "TGFB1",
        "ARG1",
        "CCL18",
        "MAFB",
        "VSIG4",
        "FOLR2",
        "MARCO",
    ],
    "microglia_interferon": ["ISG15", "IFIT1", "IFIT2", "IFIT3", "MX1", "OAS1", "OAS2", "GBP1", "STAT1", "IRF7"],
    "microglia_phagocytic_lipid": [
        "APOE",
        "APOC1",
        "LPL",
        "SPP1",
        "GPNMB",
        "TREM2",
        "CTSD",
        "CTSB",
        "LGALS3",
        "FABP5",
        "ABCA1",
    ],
    "microglia_antigen_presentation": ["HLA-DRA", "HLA-DRB1", "HLA-DPA1", "HLA-DPB1", "CD74", "CIITA", "B2M", "TAP1"],
    "microglia_hypoxia_stress": ["HIF1A", "VEGFA", "CA9", "LDHA", "BNIP3", "NDRG1", "SLC2A1", "HMOX1"],
    "microglia_complement": ["C1QA", "C1QB", "C1QC", "C3", "CFB", "SERPING1"],
    "microglia_proliferation": ["MKI67", "TOP2A", "CENPF", "UBE2C", "TYMS", "PCNA"],
}


def parse_genes(sentence: str) -> list[str]:
    return [gene.upper() for gene in sentence.split() if gene.strip()]


def score_marker_set(genes: list[str], markers: list[str]) -> dict[str, object]:
    rank = {gene: idx + 1 for idx, gene in enumerate(genes)}
    marker_upper = [marker.upper() for marker in markers]
    hits = [(marker, rank[marker]) for marker in marker_upper if marker in rank]
    weighted = sum((1001 - marker_rank) / 1000 for _, marker_rank in hits) / len(marker_upper)
    return {
        "n_markers": len(marker_upper),
        "n_hits": len(hits),
        "frac_hits": len(hits) / len(marker_upper),
        "weighted": weighted,
        "top100": sum(1 for _, marker_rank in hits if marker_rank <= 100),
        "top200": sum(1 for _, marker_rank in hits if marker_rank <= 200),
        "top500": sum(1 for _, marker_rank in hits if marker_rank <= 500),
        "hits": sorted(hits, key=lambda item: item[1]),
    }


def load_inputs() -> tuple[dict[str, dict[str, str]], dict[str, dict[str, str]]]:
    metadata = {}
    with META.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f, dialect="excel-tab")
        for row in reader:
            if row["Final_Annotation"] in {"Astrocytes", "Microglia"}:
                metadata[row["Cell"]] = row

    predictions = {}
    with PRED.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row["cell_name"] in metadata:
                predictions[row["cell_name"]] = row

    missing = sorted(set(metadata) - set(predictions))
    if missing:
        raise RuntimeError(f"Missing Cell2Sentence predictions for metadata cells: {missing[:10]}")
    return metadata, predictions


def analyze_group(
    group_name: str,
    cells: list[str],
    metadata: dict[str, dict[str, str]],
    predictions: dict[str, dict[str, str]],
    marker_sets: dict[str, list[str]],
) -> tuple[list[dict[str, object]], dict[str, dict[str, float]], list[dict[str, object]]]:
    per_cell = []
    gene_counts = Counter()
    ranks_by_gene = defaultdict(list)

    for cell in cells:
        genes = parse_genes(predictions[cell]["cell_sentence"])
        for idx, gene in enumerate(genes, start=1):
            gene_counts[gene] += 1
            ranks_by_gene[gene].append(idx)

        row: dict[str, object] = {
            "Cell": cell,
            "Final_Annotation": metadata[cell]["Final_Annotation"],
            "ID": metadata[cell]["ID"],
            "subtype": metadata[cell]["subtype"],
            "C2S_predicted_cell_type": predictions[cell]["predicted_cell_type"],
        }
        for set_name, markers in marker_sets.items():
            score = score_marker_set(genes, markers)
            row[f"{set_name}_weighted"] = round(float(score["weighted"]), 5)
            row[f"{set_name}_hits"] = score["n_hits"]
            row[f"{set_name}_top100"] = score["top100"]
            row[f"{set_name}_top200"] = score["top200"]
            row[f"{set_name}_hit_genes"] = ";".join(f"{gene}:{rank}" for gene, rank in score["hits"])
        per_cell.append(row)

    summary = {}
    for set_name in marker_sets:
        values = [float(row[f"{set_name}_weighted"]) for row in per_cell]
        hits = [int(row[f"{set_name}_hits"]) for row in per_cell]
        top100 = [int(row[f"{set_name}_top100"]) for row in per_cell]
        summary[set_name] = {
            "mean_weighted": round(statistics.mean(values), 5),
            "median_weighted": round(statistics.median(values), 5),
            "mean_hits": round(statistics.mean(hits), 3),
            "median_hits": round(statistics.median(hits), 3),
            "cells_with_any_hit_pct": round(100 * sum(hit > 0 for hit in hits) / len(hits), 1),
            "mean_top100_hits": round(statistics.mean(top100), 3),
        }

    marker_freq = []
    for set_name, markers in marker_sets.items():
        for marker in [m.upper() for m in markers]:
            ranks = ranks_by_gene.get(marker, [])
            if not ranks:
                continue
            marker_freq.append(
                {
                    "group": group_name,
                    "marker_set": set_name,
                    "gene": marker,
                    "cells_present": len(ranks),
                    "pct_cells": round(100 * len(ranks) / len(cells), 1),
                    "median_rank": round(statistics.median(ranks), 1),
                    "top100_cells": sum(rank <= 100 for rank in ranks),
                    "top200_cells": sum(rank <= 200 for rank in ranks),
                }
            )
    marker_freq.sort(key=lambda row: (row["group"], row["marker_set"], -row["pct_cells"], row["median_rank"]))
    return per_cell, summary, marker_freq


def write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    if not rows:
        return
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def add_sheet(wb: Workbook, title: str, rows: list[dict[str, object]]) -> None:
    ws = wb.create_sheet(title=title)
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row in rows:
        ws.append([row[header] for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, header in enumerate(headers, start=1):
        width = 18
        if header.endswith("_hit_genes"):
            width = 46
        elif header in {"Cell", "marker_set"}:
            width = 32
        elif header == "gene":
            width = 14
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width


def write_workbook(
    astro_rows: list[dict[str, object]],
    micro_rows: list[dict[str, object]],
    freq_rows: list[dict[str, object]],
    summary_rows: list[dict[str, object]],
) -> Path:
    path = OUTDIR / "GEO_metadata_astrocyte_microglia_rank_marker_analysis.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Interpretation"
    ws.append(["Analysis", "Result"])
    ws.append(
        [
            "Cohorts",
            "GEO metadata Final_Annotation labels were used: 47 Astrocytes and 288 Microglia. Cell2Sentence 1000-gene rankings were used only for marker scoring.",
        ]
    )
    ws.append(
        [
            "Scoring",
            "Each marker set was scored per cell as the average rank weight across all markers: present marker weight = (1001 - rank) / 1000; absent marker weight = 0.",
        ]
    )
    ws.append(
        [
            "Interpretation",
            "Higher values indicate that more markers from a program are present and/or occur earlier in the 1000-gene cell_sentence.",
        ]
    )
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 120
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    add_sheet(wb, "Summary Scores", summary_rows)
    add_sheet(wb, "Marker Frequencies", freq_rows)
    add_sheet(wb, "Astrocyte Cell Scores", astro_rows)
    add_sheet(wb, "Microglia Cell Scores", micro_rows)
    wb.save(path)

    check_wb = load_workbook(path, read_only=True)
    if check_wb["Astrocyte Cell Scores"].max_row - 1 != len(astro_rows):
        raise RuntimeError("Astrocyte workbook row-count check failed")
    if check_wb["Microglia Cell Scores"].max_row - 1 != len(micro_rows):
        raise RuntimeError("Microglia workbook row-count check failed")
    return path


def make_figure(summary: dict[str, dict[str, dict[str, float]]], freq_rows: list[dict[str, object]]) -> tuple[Path, Path]:
    MPL_CACHE.mkdir(exist_ok=True)
    os.environ.setdefault("MPLCONFIGDIR", str(MPL_CACHE))
    os.environ.setdefault("XDG_CACHE_HOME", str(MPL_CACHE))
    import matplotlib.pyplot as plt
    import numpy as np

    plt.rcParams.update(
        {
            "text.color": "black",
            "axes.labelcolor": "black",
            "axes.titlecolor": "black",
            "xtick.color": "black",
            "ytick.color": "black",
        }
    )

    def top_freq(group: str, marker_set: str, n: int = 8) -> list[dict[str, object]]:
        rows = [row for row in freq_rows if row["group"] == group and row["marker_set"] == marker_set]
        rows.sort(key=lambda row: (-float(row["pct_cells"]), float(row["median_rank"])))
        return rows[:n]

    def rank_gradient_colors(cmap_name: str, ranks: list[float]) -> list[tuple[float, float, float, float]]:
        cmap = plt.get_cmap(cmap_name)
        if not ranks:
            return []
        min_rank = min(ranks)
        max_rank = max(ranks)
        if min_rank == max_rank:
            return [cmap(0.88) for _ in ranks]
        colors = []
        for rank in ranks:
            # Lower median rank means the marker is more prominent, so it gets
            # the darker end of the colormap.
            prominence = 1 - ((rank - min_rank) / (max_rank - min_rank))
            colors.append(cmap(0.25 + prominence * 0.65))
        return colors

    def set_black_text(ax):
        ax.title.set_color("black")
        ax.xaxis.label.set_color("black")
        ax.yaxis.label.set_color("black")
        for tick in ax.get_xticklabels() + ax.get_yticklabels():
            tick.set_color("black")

    def score_bar(ax, labels, values, title, colors):
        x = np.arange(len(labels))
        bars = ax.bar(x, values, color=colors, edgecolor="#222222", linewidth=0.7)
        ax.set_xticks(x)
        ax.set_xticklabels(labels, rotation=20, ha="right", fontsize=8)
        ax.set_ylabel("Mean rank-weighted score")
        ax.set_title(title, loc="left", fontweight="bold")
        ax.set_ylim(0, max(values) * 1.35 if max(values) else 0.1)
        ax.spines[["top", "right"]].set_visible(False)
        ax.spines[["bottom", "left"]].set_color("black")
        ax.tick_params(colors="black")
        for bar, value in zip(bars, values):
            ax.text(bar.get_x() + bar.get_width() / 2, value + max(values) * 0.03, f"{value:.3f}", ha="center", fontsize=8, color="black")
        set_black_text(ax)

    def freq_bar(ax, rows, title, cmap_name):
        genes = [str(row["gene"]) for row in rows][::-1]
        pct = [float(row["pct_cells"]) for row in rows][::-1]
        ranks = [float(row["median_rank"]) for row in rows][::-1]
        y = np.arange(len(genes))
        colors = rank_gradient_colors(cmap_name, ranks)
        ax.barh(y, pct, color=colors, edgecolor="#222222", linewidth=0.7)
        ax.set_yticks(y)
        ax.set_yticklabels(genes, fontstyle="italic", fontsize=8)
        ax.set_xlabel("Cells containing marker (%)")
        ax.set_title(title, loc="left", fontweight="bold")
        ax.set_xlim(0, max(pct) * 1.25 if pct else 1)
        ax.spines[["top", "right"]].set_visible(False)
        ax.spines[["bottom", "left"]].set_color("black")
        ax.tick_params(colors="black")
        for yi, p, rank in zip(y, pct, ranks):
            ax.text(p + 1, yi, f"{p:.1f}% | r={rank:.0f}", va="center", fontsize=8, color="black")
        set_black_text(ax)

    fig = plt.figure(figsize=(12, 9), constrained_layout=True)
    grid = fig.add_gridspec(2, 2, height_ratios=[1, 1.3])
    ax1 = fig.add_subplot(grid[0, 0])
    ax2 = fig.add_subplot(grid[0, 1])
    ax3 = fig.add_subplot(grid[1, 0])
    ax4 = fig.add_subplot(grid[1, 1])

    astro_labels = ["Homeostatic", "Reactive/\nactivated", "Inflammatory/\nneurotoxic", "Reparative/\ntrophic"]
    astro_values = [
        summary["Astrocytes"]["astro_homeostatic"]["mean_weighted"],
        summary["Astrocytes"]["astro_reactive_active"]["mean_weighted"],
        summary["Astrocytes"]["astro_inflammatory_neurotoxic"]["mean_weighted"],
        summary["Astrocytes"]["astro_reparative_trophic"]["mean_weighted"],
    ]
    score_bar(ax1, astro_labels, astro_values, "A. GEO metadata-labeled astrocyte states (n=47)", ["#577590", "#F3722C", "#C1121F", "#43AA8B"])

    micro_labels = ["Homeostatic", "M1\ninflam.", "M2\nimmunoreg.", "Interferon", "Phagocytic/\nlipid", "Antigen\npresentation", "Hypoxia/\nstress", "Complement", "Prolif."]
    micro_keys = [
        "microglia_homeostatic",
        "microglia_m1_inflammatory",
        "microglia_m2_immunoregulatory",
        "microglia_interferon",
        "microglia_phagocytic_lipid",
        "microglia_antigen_presentation",
        "microglia_hypoxia_stress",
        "microglia_complement",
        "microglia_proliferation",
    ]
    micro_values = [summary["Microglia"][key]["mean_weighted"] for key in micro_keys]
    score_bar(ax2, micro_labels, micro_values, "B. GEO metadata-labeled microglia programs (n=288)", ["#577590", "#D62828", "#2A9D8F", "#7B2CBF", "#F4A261", "#457B9D", "#6C757D", "#8D99AE", "#E76F51"])

    freq_bar(ax3, top_freq("Astrocytes", "astro_reactive_active"), "C. Reactive/activated astrocyte markers", "Oranges")
    freq_bar(ax4, top_freq("Microglia", "microglia_phagocytic_lipid"), "D. Phagocytic/lipid microglia markers", "YlGnBu")
    fig.suptitle("Rank-based marker analysis of GEO metadata-labeled astrocytes and microglia", fontsize=14, fontweight="bold", color="black")
    fig.supxlabel("Marker scores use Cell2Sentence 1000-gene ranks; lower r indicates higher marker prominence.", fontsize=9, color="black")

    png = OUTDIR / "GEO_metadata_astrocyte_microglia_rank_marker_figure.png"
    pdf = OUTDIR / "GEO_metadata_astrocyte_microglia_rank_marker_figure.pdf"
    fig.savefig(png, dpi=300, bbox_inches="tight")
    fig.savefig(pdf, bbox_inches="tight")
    plt.close(fig)
    return png, pdf


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Calculate supplementary rank-weighted Cell2Sentence state-program scores.")
    parser.add_argument("--metadata", type=Path, required=True, help="GSE283839 metadata TSV.")
    parser.add_argument("--predictions", type=Path, required=True, help="Cell2Sentence predictions CSV containing cell_name, cell_sentence, and predicted_cell_type.")
    parser.add_argument("--output-dir", type=Path, required=True)
    return parser.parse_args()


def main() -> None:
    global OUTDIR, META, PRED, MPL_CACHE
    args = parse_args()
    META = args.metadata.resolve()
    PRED = args.predictions.resolve()
    OUTDIR = args.output_dir.resolve()
    MPL_CACHE = OUTDIR / "matplotlib_cache"
    for path in (META, PRED):
        if not path.is_file():
            raise FileNotFoundError(path)
    OUTDIR.mkdir(parents=True, exist_ok=True)
    metadata, predictions = load_inputs()
    astro_cells = sorted(cell for cell, row in metadata.items() if row["Final_Annotation"] == "Astrocytes")
    micro_cells = sorted(cell for cell, row in metadata.items() if row["Final_Annotation"] == "Microglia")

    astro_rows, astro_summary, astro_freq = analyze_group("Astrocytes", astro_cells, metadata, predictions, ASTRO_MARKERS)
    micro_rows, micro_summary, micro_freq = analyze_group("Microglia", micro_cells, metadata, predictions, MICROGLIA_MARKERS)

    for row in astro_rows:
        row["reactive_minus_homeostatic"] = round(row["astro_reactive_active_weighted"] - row["astro_homeostatic_weighted"], 5)
    for row in micro_rows:
        row["m2_minus_m1"] = round(row["microglia_m2_immunoregulatory_weighted"] - row["microglia_m1_inflammatory_weighted"], 5)
        program_cols = [f"{key}_weighted" for key in MICROGLIA_MARKERS]
        row["dominant_rank_program"] = max(program_cols, key=lambda col: row[col]).replace("_weighted", "")

    summary = {"Astrocytes": astro_summary, "Microglia": micro_summary}
    freq_rows = astro_freq + micro_freq
    summary_rows = []
    for group_name, group_summary in summary.items():
        for marker_set, values in group_summary.items():
            summary_rows.append({"group": group_name, "marker_set": marker_set, **values})

    report = {
        "n_astrocytes": len(astro_rows),
        "n_microglia": len(micro_rows),
        "summary": summary,
        "astro_reactive_minus_homeostatic": {
            "mean": round(statistics.mean(row["reactive_minus_homeostatic"] for row in astro_rows), 5),
            "median": round(statistics.median(row["reactive_minus_homeostatic"] for row in astro_rows), 5),
            "pct_reactive_gt_homeostatic": round(100 * sum(row["reactive_minus_homeostatic"] > 0 for row in astro_rows) / len(astro_rows), 1),
        },
        "microglia_m2_minus_m1": {
            "mean": round(statistics.mean(row["m2_minus_m1"] for row in micro_rows), 5),
            "median": round(statistics.median(row["m2_minus_m1"] for row in micro_rows), 5),
            "pct_m2_gt_m1": round(100 * sum(row["m2_minus_m1"] > 0 for row in micro_rows) / len(micro_rows), 1),
            "pct_m1_gt_m2": round(100 * sum(row["m2_minus_m1"] < 0 for row in micro_rows) / len(micro_rows), 1),
        },
        "microglia_dominant_rank_program_counts": dict(Counter(row["dominant_rank_program"] for row in micro_rows).most_common()),
    }

    write_csv(OUTDIR / "GEO_metadata_astrocyte_rank_marker_scores.csv", astro_rows)
    write_csv(OUTDIR / "GEO_metadata_microglia_rank_marker_scores.csv", micro_rows)
    write_csv(OUTDIR / "GEO_metadata_astrocyte_microglia_marker_frequencies.csv", freq_rows)
    write_csv(OUTDIR / "GEO_metadata_astrocyte_microglia_summary_scores.csv", summary_rows)
    (OUTDIR / "GEO_metadata_astrocyte_microglia_rank_marker_report.json").write_text(json.dumps(report, indent=2), encoding="utf-8")
    workbook = write_workbook(astro_rows, micro_rows, freq_rows, summary_rows)
    png, pdf = make_figure(summary, freq_rows)

    print(json.dumps(report, indent=2))
    print(workbook)
    print(png)
    print(pdf)


if __name__ == "__main__":
    main()
